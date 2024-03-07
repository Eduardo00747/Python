# Importando as bibliotecas necessárias
import os
import requests
from bs4 import BeautifulSoup
import openpyxl

# Função para obter as notícias do site da Globo
def get_news():
    # URL do site da Globo
    url = 'https://www.globo.com'

    # Fazendo uma requisição para a página
    page = requests.get(url)
    
    # Criando uma sopa de HTML usando BeautifulSoup
    soup = BeautifulSoup(page.text, 'html.parser')
    
    # Encontrando todos os elementos 'a' que contêm links
    noticias = soup.find_all('a')

    # Classes específicas para identificar as notícias
    tgt_class1 = 'post__title'
    tgt_class2 = 'post-multicontent__link--title__text'

    # Dicionário para armazenar as notícias (título: link)
    news_dict = {}
    for noticia in noticias:
        # Verificando se a notícia tem um elemento h2 e a classe desejada
        if (noticia.h2 is not None) and (noticia.h2.get('class') is not None):
            # Adicionando a notícia ao dicionário com o título como chave e o link como valor
            if tgt_class1 in noticia.h2.get('class'):
                news_dict[noticia.h2.text] = noticia.get('href')
            if tgt_class2 in noticia.h2.get('class'):
                news_dict[noticia.h2.text] = noticia.get('href')
    return news_dict

# Função para exibir as notícias no terminal
def display_news(news):
    for idx, (title, link) in enumerate(news.items(), start=1):
        print(f"{idx}. {title}")
        print(f"   Link: {link}")
        print()

# Função para criar um arquivo Excel com as notícias
def create_excel(news):
    # Criando uma nova planilha do Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Noticias"

    # Adicionando cabeçalhos à planilha
    ws['A1'] = "Título"
    ws['B1'] = "Link"

    # Preenchendo a planilha com os títulos e links das notícias
    for idx, (title, link) in enumerate(news.items(), start=2):
        ws[f'A{idx}'] = title
        ws[f'B{idx}'] = link

    # Obtendo o caminho absoluto para a pasta do script
    script_path = os.path.dirname(os.path.abspath(__file__))
    # Criando o caminho completo para o arquivo Excel
    excel_filename = os.path.join(script_path, "noticias.xlsx")

    # Salvando a planilha do Excel
    wb.save(excel_filename)
    print(f"Arquivo Excel '{excel_filename}' criado com sucesso.")

# Bloco principal de execução
if __name__ == "__main__":
    # Obtendo as notícias
    news = get_news()
    
    # Verificando se há notícias
    if news:
        # Exibindo as notícias no terminal
        print(f"Encontradas {len(news)} notícias:")
        display_news(news)
        # Criando o arquivo Excel
        create_excel(news)
    else:
        # Mensagem se nenhuma notícia for encontrada
        print("Nenhuma notícia encontrada.")
